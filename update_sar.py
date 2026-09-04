#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
update_sar.py — Processador ETL para o Dashboard SAR do BI JLE Telecom
Lê a planilha oficial do SAR (Google Sheets em nuvem com fallback em rede/cache local)
e compila os dados normalizados em sar_data.js.
"""

import sys
import os
import io
import csv
import json
import datetime
import urllib.request
import unicodedata
import re
import openpyxl

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

GOOGLE_SHEET_ID = "1kQyIsIDmsnunTbHU46n_3FmeL8ddbGGHnXHo6FXAfq4"
GOOGLE_SHEET_GID = "1221770117"
GOOGLE_SHEET_CSV_URL = f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/export?format=csv&gid={GOOGLE_SHEET_GID}"

# Meses em português
MESES_PT = [
    "", "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
    "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"
]

def parse_date(val):
    """Converte valor para string ISO 'YYYY-MM-DD'."""
    if val is None:
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    val_str = str(val).strip()
    if not val_str or val_str in ("-", "None", "none", "null", "N/A", "n/a", ""):
        return None
    
    # Formato DD/MM/YYYY
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', val_str)
    if m:
        try:
            return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
        except Exception:
            pass

    # Formato YYYY-MM-DD
    m = re.match(r'^(\d{4})-(\d{1,2})-(\d{1,2})', val_str)
    if m:
        try:
            return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        except Exception:
            pass

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            dt = datetime.datetime.strptime(val_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None

def format_date_br(iso_str):
    """Converte 'YYYY-MM-DD' para 'DD/MM/YYYY'."""
    if not iso_str or len(iso_str) < 10:
        return "-"
    try:
        parts = iso_str[:10].split("-")
        if len(parts) == 3:
            return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except Exception:
        pass
    return str(iso_str)

def get_competencia(iso_date):
    """Retorna 'MÊS/ANO' a partir de 'YYYY-MM-DD'."""
    if not iso_date:
        return "NÃO INFORMADO"
    try:
        parts = iso_date.split("-")
        ano = parts[0]
        mes_num = int(parts[1])
        if 1 <= mes_num <= 12:
            return f"{MESES_PT[mes_num]}/{ano}"
    except Exception:
        pass
    return "NÃO INFORMADO"

def to_number(val):
    """Converte com segurança para float/int ou 0 tratando formatos numéricos pt-BR e moedas (R$)."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "-", "n/a", "#error!", "#ref!", "#value!", "#n/d"):
        return 0.0
    try:
        # Remover R$, espaços normais e non-breaking spaces
        s = val_str.replace("R$", "").replace("r$", "").replace("\xa0", "").replace(" ", "").strip()
        # Se contiver vírgula como decimal (ex: 1.121,35 ou 596,00)
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        return round(float(s), 2)
    except Exception:
        return 0.0

def norm_h(text):
    """Normaliza texto de cabeçalho removendo acentos, pontuações e símbolos."""
    if not text:
        return ""
    s = str(text).strip().upper()
    s = s.replace('º', ' ').replace('ª', ' ').replace('°', ' ').replace('.', ' ')
    s = "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', ' ', s).strip()

def clean_str(val):
    """Limpa e padroniza string."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("none", "null", "-", "n/a") else s

def load_from_google_sheets(url):
    """Baixa e converte dados da planilha Google Sheets via export CSV."""
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=12) as resp:
            content = resp.read().decode('utf-8')
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)
            if rows and len(rows) > 3:
                return rows
    except Exception as e:
        print(f"Aviso ao acessar Google Sheets ({url}): {e}")
    return None

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    rows = None
    source_desc = ""

    # 1. Tentar ler diretamente do Google Sheets
    print(f"Tentando sincronizar base SAR online do Google Sheets...")
    gs_url = os.environ.get("GOOGLE_SHEET_SAR_URL") or GOOGLE_SHEET_CSV_URL
    rows = load_from_google_sheets(gs_url)
    
    if rows:
        source_desc = f"Google Sheets ({GOOGLE_SHEET_ID})"
        print(f"[OK] Sucesso! {len(rows)} linhas baixadas do Google Sheets.")
        # Salvar cópia local de contingência em CSV
        try:
            cache_csv_path = os.path.join(base_dir, "sar_local.csv")
            with open(cache_csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.writer(f)
                writer.writerows(rows)
        except Exception:
            pass

    # 2. Fallback para arquivos locais (se Google Sheets falhar ou offline)
    if not rows:
        print("Buscando fontes locais / cache de contingência...")
        input_file = None
        if len(sys.argv) > 1 and os.path.exists(sys.argv[1]):
            input_file = sys.argv[1]
        else:
            candidates = [
                os.path.join(base_dir, "sar_local.csv"),
                os.path.join(base_dir, "Planilha_Operacional_SAR_JLE.xlsx"),
                os.path.join(base_dir, "sar_temp.xlsx"),
                os.path.join(base_dir, "sar_local.xlsx")
            ]
            for c in candidates:
                if os.path.exists(c):
                    input_file = c
                    break

        if input_file and input_file.endswith(".csv"):
            with open(input_file, "r", encoding="utf-8") as f:
                rows = list(csv.reader(f))
            source_desc = f"Cache CSV Local ({os.path.basename(input_file)})"
        elif input_file and input_file.endswith(".xlsx"):
            wb = openpyxl.load_workbook(input_file, data_only=True, read_only=True)
            sheet_name = "SAR Operacional" if "SAR Operacional" in wb.sheetnames else ("Ext. MDU" if "Ext. MDU" in wb.sheetnames else wb.sheetnames[0])
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            source_desc = f"Arquivo Excel ({os.path.basename(input_file)})"

    if not rows or len(rows) < 2:
        print("ERRO CRÍTICO: Não foi possível carregar dados do SAR.", file=sys.stderr)
        sys.exit(1)

    print(f"Processando dados da fonte: {source_desc}")

    # Localizar a linha de cabeçalho
    header_row_idx = None
    for r_idx, row in enumerate(rows[:15]):
        row_norm = [norm_h(x) for x in row if x is not None]
        if any(x.startswith("ETAPA") for x in row_norm):
            continue
        if any("COD" in x for x in row_norm) and any("CIDADE" in x or "AREA" in x or "ENTRADA" in x for x in row_norm):
            header_row_idx = r_idx
            print(f"Linha de cabeçalho detectada no índice {header_row_idx + 1}")
            break

    if header_row_idx is None:
        header_row_idx = 0 if len(rows[0]) > 10 else 2
        print(f"Usando cabeçalho padrão índice {header_row_idx + 1}")

    # Construir mapa de cabeçalho dinâmico
    header_map = {}
    header_cells = rows[header_row_idx]
    for col_idx, cell in enumerate(header_cells):
        if cell:
            n_key = norm_h(cell)
            header_map[n_key] = col_idx

    print(f"Colunas mapeadas dinamicamente: {len(header_map)}")

    def get_col(row, *aliases, default_idx=None):
        for a in aliases:
            n = norm_h(a)
            if n in header_map:
                idx = header_map[n]
                return row[idx] if idx < len(row) else None
        if default_idx is not None and default_idx < len(row):
            return row[default_idx]
        return None

    records = []
    today = datetime.date.today()

    for r_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        if not row or len(row) < 2:
            continue

        cod = clean_str(get_col(row, "CODIGO SAR", "CODIGO", "COD.", "COD", default_idx=0))
        cidade = clean_str(get_col(row, "CIDADE", "MUNICIPIO", default_idx=4))

        # REGRA FUNDAMENTAL: Descartar linhas de rascunho onde a Cidade está vazia
        if not cod or not cidade:
            continue

        area_tecnica = clean_str(get_col(row, "AREA TECNICA", "AREA", "AT", default_idx=1))
        node = clean_str(get_col(row, "NODE", "NO", default_idx=2))
        site = clean_str(get_col(row, "SITE", "LOCAL", default_idx=3))
        condominio = clean_str(get_col(row, "CONDOMINIO", "COMDOMINIO", "CLIENTE", "PREDIO", default_idx=5))
        endereco = clean_str(get_col(row, "ENDERECO", "LOGRADOURO", "RUA", default_idx=6))
        caixa_mdu = clean_str(get_col(row, "CAIXA MDU", "CX MDU", "CAIXA", default_idx=7))
        servico = clean_str(get_col(row, "SERVICO / ESCOPO", "SERVICO", "ESCOPO", default_idx=8))
        classe_l = clean_str(get_col(row, "EXECUTOR LINHA (CLASSE L)", "EXECUTOR (CLASSE L)", "EXECUTOR LINHA", "CLASSE L", default_idx=9))
        classe_f = clean_str(get_col(row, "EXECUTOR FUSAO (CLASSE F)", "EXECUTOR (CLASSE F)", "EXECUTOR FUSAO", "CLASSE F", default_idx=10))
        
        # Campos de Campo / Projeto
        projetado = clean_str(get_col(row, "PROJETADO", default_idx=11))
        executado = clean_str(get_col(row, "EXECUTADO", default_idx=12))
        observacao_op = clean_str(get_col(row, "OBSERVACAO", "OBSERVAÇÃO", "SITUACAO", default_idx=13))
        
        # Datas de Operação
        dt_entrada_iso = parse_date(get_col(row, "DATA ENTRADA", "DATA DE ENTRADA", "ENTRADA", default_idx=14))
        dt_inicio_iso = parse_date(get_col(row, "INICIO EM", "INÍCIO EM", "DATA INICIO", default_idx=15))
        dt_previsao_iso = parse_date(get_col(row, "PREVISAO", "PREVISÃO", "PREVISAO PARA", default_idx=16))
        dt_entrega_iso = parse_date(get_col(row, "DATA ENTREGA", "DATA DE ENTREGA", "ENTREGA", default_idx=17))
        relatorio_ppt = clean_str(get_col(row, "RELATORIO PPT", "RELATÓRIO PPT", "RELATORIO PPT / FOTOS", default_idx=18))
        data_envio_med = clean_str(get_col(row, "DATA ENVIO MEDICAO", "DATA ENVIO MEDIÇÃO", "MTA ENVIO MEDICAO", "ENVIO MEDICAO", default_idx=19))

        # Tempo, Prazo e Atraso
        tempo_raw = get_col(row, "TEMPO (DIAS)", "TEMPO DIAS", "TEMPO", default_idx=17)
        prazo_raw = clean_str(get_col(row, "PRAZO (SLA 3 DIAS)", "PRAZO SLA 3 DIAS", "PRAZO", "SLA", default_idx=18)).upper()
        atraso_raw = get_col(row, "ATRASO (DIAS)", "ATRASO DIAS", "ATRASO", default_idx=19)

        # Status Geral SAR - Exatamente o nome da Planilha Google Sheets
        status_geral_raw = clean_str(get_col(row, "STATUS STATUS GERAL SAR", "STATUS GERAL SAR", "STATUS GERAL", "STATUS", default_idx=21))
        status = status_geral_raw.strip() if status_geral_raw else "EM ANDAMENTO"

        # Colunas de Medição & Faturamento (Douglas)
        dt_medicao_iso = parse_date(get_col(row, "ETAPA 2: MEDICAO (PREENCHIMENTO: DOUGLAS) DATA MEDICAO", "DATA MEDICAO", "DATA MEDIÇÃO", default_idx=35))
        valor_medicao = to_number(get_col(row, "VALOR MEDICAO (R$)", "VALOR MEDICAO", "VALOR", default_idx=36))
        previa_medicao = to_number(get_col(row, "ETAPA 2: MEDICAO (PREENCHIMENTO: DOUGLAS) PREVIA MEDICAO", "ETAPA 2: MEDICAO (PREENCHIMENTO: DOUGLAS) PRÉVIA MEDIÇÃO", "PREVIA MEDICAO", "PRÉVIA MEDIÇÃO", default_idx=34))
        total_terceiros = to_number(get_col(row, "TOTAL TERCEIROS", "TOTAL TERCEIRO", "TERCEIROS TOTAL", default_idx=33))

        # Quantidades dos Itens de LPU de Terceiros
        # Classe L (Cabos e Cordoalha: 2.11, 2.12, 2.15, 1.13)
        q_211 = to_number(get_col(row, "TERCEIROS 2.11 CB AS", "2.11 CB AS", "2.11", default_idx=22))
        q_212 = to_number(get_col(row, "2.12 CB SUB", "2.12", default_idx=23))
        q_215 = to_number(get_col(row, "2.15 CB ESP", "2.15", default_idx=24))
        q_113 = to_number(get_col(row, "1.13 CORD", "1.13", default_idx=25))

        # Classe F (Caixas, Fusão, Testes e Outros: 3.11, 3.17, 3.18, 3.15, 3.13, 3.14, 3.12)
        q_311 = to_number(get_col(row, "3.11 CX EM", "3.11", default_idx=26))
        q_317 = to_number(get_col(row, "3.17 DIO/DGO", "3.17", default_idx=27))
        q_318 = to_number(get_col(row, "3.18 NAP/CTO", "3.18", default_idx=28))
        q_315 = to_number(get_col(row, "3.15 AB/FE", "3.15 AB/FC", "3.15", default_idx=29))
        q_313 = to_number(get_col(row, "3.13 FUS/EME", "3.13", default_idx=30))
        q_314 = to_number(get_col(row, "3.14 OTDR", "3.14", default_idx=31))
        q_312 = to_number(get_col(row, "3.12 DER/INS", "3.12 DER/ANS", "3.12", default_idx=32))

        # Cálculo teórico de rateio LPU
        calc_l = (q_211 * 1.20) + (q_212 * 1.00) + (q_215 * 1.45) + (q_113 * 0.75)
        calc_f = (q_311 * 65.00) + (q_317 * 50.00) + (q_318 * 60.00) + (q_315 * 80.00) + (q_313 * 9.00) + (q_314 * 3.50) + (q_312 * 15.00)

        val_l = 0.0
        val_f = 0.0
        if total_terceiros > 0:
            if classe_l and classe_f:
                if calc_l > 0 and calc_f > 0:
                    val_l = round(total_terceiros * (calc_l / (calc_l + calc_f)), 2)
                    val_f = round(total_terceiros - val_l, 2)
                elif calc_l > 0:
                    val_l = min(total_terceiros, round(calc_l, 2))
                    val_f = round(total_terceiros - val_l, 2)
                elif calc_f > 0:
                    val_f = min(total_terceiros, round(calc_f, 2))
                    val_l = round(total_terceiros - val_f, 2)
                else:
                    val_l = round(total_terceiros / 2.0, 2)
                    val_f = round(total_terceiros - val_l, 2)
            elif classe_l and not classe_f:
                val_l = total_terceiros
                val_f = 0.0
            elif classe_f and not classe_l:
                val_f = total_terceiros
                val_l = 0.0
            else:
                if calc_l > 0 and calc_f > 0:
                    val_l = round(total_terceiros * (calc_l / (calc_l + calc_f)), 2)
                    val_f = round(total_terceiros - val_l, 2)
                elif calc_l > 0:
                    val_l = min(total_terceiros, round(calc_l, 2))
                    val_f = round(total_terceiros - val_l, 2)
                else:
                    val_f = total_terceiros
                    val_l = 0.0
        else:
            val_l = round(calc_l, 2)
            val_f = round(calc_f, 2)
            total_terceiros = round(val_l + val_f, 2)

        num_wf = clean_str(get_col(row, "N WF", "NO WF", "NUM WF", "Nº WF", "WORKFLOW", default_idx=37))
        dt_pedido_iso = parse_date(get_col(row, "DATA PEDIDO", default_idx=38))
        num_pedido = clean_str(get_col(row, "N DO PEDIDO", "NO DO PEDIDO", "Nº DO PEDIDO", "PEDIDO", default_idx=39))
        observacoes = clean_str(get_col(row, "OBSERVACOES", "OBSERVAÇÕES", "OBS", default_idx=40))
        status_wf = "100% - OK" if "IMPLANTADO" in status or "APROV" in status else ""

        # Cálculo do Tempo e SLA em dias úteis
        tempo_dias = to_number(tempo_raw)
        if tempo_dias <= 0 and dt_entrada_iso:
            try:
                import numpy as np
                d1 = datetime.datetime.strptime(dt_entrada_iso[:10], "%Y-%m-%d").date()
                d2 = None
                if dt_medicao_iso:
                    d2 = datetime.datetime.strptime(dt_medicao_iso[:10], "%Y-%m-%d").date()
                elif dt_entrega_iso:
                    d2 = datetime.datetime.strptime(dt_entrega_iso[:10], "%Y-%m-%d").date()
                elif "CONCLU" not in status.upper() and "CANCEL" not in status.upper():
                    d2 = today

                if d1 and d2:
                    if d2 >= d1:
                        tempo_dias = int(np.busday_count(d1, d2))
                    else:
                        tempo_dias = 0
            except Exception:
                pass

        if "NO PRAZO" in prazo_raw or "DENTRO" in prazo_raw:
            prazo = "NO PRAZO"
        elif "ATRASAD" in prazo_raw or "FORA" in prazo_raw:
            prazo = "ATRASADO"
        else:
            prazo = "ATRASADO" if tempo_dias > 3 else "NO PRAZO"

        atraso_dias = to_number(atraso_raw)
        if prazo == "ATRASADO" and atraso_dias <= 0 and tempo_dias > 3:
            atraso_dias = tempo_dias - 3
        elif prazo == "NO PRAZO":
            atraso_dias = 0

        # Competência e Períodos (Data de Entrada)
        competencia = get_competencia(dt_entrada_iso)
        ano_entrada = dt_entrada_iso[:4] if dt_entrada_iso else "NÃO INFORMADO"
        mes_num = dt_entrada_iso[5:7] if dt_entrada_iso and len(dt_entrada_iso) >= 7 else ""
        mes_idx = int(mes_num) if mes_num.isdigit() and 1 <= int(mes_num) <= 12 else 0
        mes_nome = MESES_PT[mes_idx] if mes_idx > 0 else "NÃO INFORMADO"

        # Competência e Períodos (Data de Entrega — Referência para Fechamento de Terceiros)
        competencia_entrega = get_competencia(dt_entrega_iso)
        ano_entrega = dt_entrega_iso[:4] if dt_entrega_iso else "NÃO INFORMADO"
        mes_num_entrega = dt_entrega_iso[5:7] if dt_entrega_iso and len(dt_entrega_iso) >= 7 else ""
        mes_idx_entrega = int(mes_num_entrega) if mes_num_entrega.isdigit() and 1 <= int(mes_num_entrega) <= 12 else 0
        mes_nome_entrega = MESES_PT[mes_idx_entrega] if mes_idx_entrega > 0 else "NÃO INFORMADO"

        # Competência e Períodos (Data de Medição — Referência para Indicador de Medição)
        competencia_medicao = get_competencia(dt_medicao_iso) if dt_medicao_iso else "Sem Data"
        ano_medicao = dt_medicao_iso[:4] if dt_medicao_iso else "SEM DATA"
        mes_num_medicao = dt_medicao_iso[5:7] if dt_medicao_iso and len(dt_medicao_iso) >= 7 else ""
        mes_idx_medicao = int(mes_num_medicao) if mes_num_medicao.isdigit() and 1 <= int(mes_num_medicao) <= 12 else 0
        mes_nome_medicao = MESES_PT[mes_idx_medicao] if mes_idx_medicao > 0 else "SEM DATA"

        # Normalização canônica do Status para fins de Medição
        st_clean = status.upper().strip()
        status_medicao_grupo = "OUTROS"
        if "MEDI" in st_clean and "ENVIAD" in st_clean:
            status_medicao_grupo = "MEDIÇÃO ENVIADA"
        elif "FINALIZ" in st_clean:
            status_medicao_grupo = "FINALIZADO"
        elif "PEDIDO" in st_clean and "EMIT" in st_clean:
            status_medicao_grupo = "PEDIDO EMITIDO"

        # Resumo formatado de itens LPU para visualização/auditoria
        itens_l_resumo = []
        if q_211 > 0: itens_l_resumo.append(f"2.11 CB AS: {q_211:g}m")
        if q_212 > 0: itens_l_resumo.append(f"2.12 CB SUB: {q_212:g}m")
        if q_215 > 0: itens_l_resumo.append(f"2.15 CB ESP: {q_215:g}m")
        if q_113 > 0: itens_l_resumo.append(f"1.13 CORD: {q_113:g}m")

        itens_f_resumo = []
        if q_311 > 0: itens_f_resumo.append(f"3.11 CX EM: {q_311:g} un")
        if q_317 > 0: itens_f_resumo.append(f"3.17 DIO/DGO: {q_317:g} un")
        if q_318 > 0: itens_f_resumo.append(f"3.18 NAP/CTO: {q_318:g} un")
        if q_315 > 0: itens_f_resumo.append(f"3.15 AB/FE: {q_315:g} un")
        if q_313 > 0: itens_f_resumo.append(f"3.13 FUS/EME: {q_313:g} un")
        if q_314 > 0: itens_f_resumo.append(f"3.14 OTDR: {q_314:g} un")
        if q_312 > 0: itens_f_resumo.append(f"3.12 DER/INS: {q_312:g} un")

        record = {
            "cod": cod,
            "area_tecnica": area_tecnica,
            "node": node,
            "site": site,
            "cidade": cidade,
            "condominio": condominio,
            "endereco": endereco,
            "caixa_mdu": caixa_mdu,
            "classe_l": classe_l,
            "classe_f": classe_f,
            "situacao": observacao_op or projetado,
            "relatorio_foto": executado,
            "servico": servico,
            "data_entrada": dt_entrada_iso,
            "data_entrada_fmt": format_date_br(dt_entrada_iso),
            "data_inicio": dt_inicio_iso,
            "data_inicio_fmt": format_date_br(dt_inicio_iso),
            "data_previsao": dt_previsao_iso,
            "data_previsao_fmt": format_date_br(dt_previsao_iso),
            "data_entrega": dt_entrega_iso,
            "data_entrega_fmt": format_date_br(dt_entrega_iso),
            "data_medicao": dt_medicao_iso,
            "data_medicao_fmt": format_date_br(dt_medicao_iso),
            "competencia_medicao": competencia_medicao,
            "ano_medicao": ano_medicao,
            "mes_medicao": mes_nome_medicao,
            "mes_num_medicao": mes_num_medicao,
            "status_medicao_grupo": status_medicao_grupo,
            "tem_medicao": bool(valor_medicao > 0),
            "num_wf": num_wf,
            "status_wf": status_wf,
            "competencia": competencia,
            "ano": ano_entrada,
            "mes": mes_nome,
            "mes_num": mes_num,
            "competencia_entrega": competencia_entrega,
            "ano_entrega": ano_entrega,
            "mes_entrega": mes_nome_entrega,
            "mes_num_entrega": mes_num_entrega,
            "status": status,
            "status_relatorio": relatorio_ppt,
            "status_medicao": data_envio_med,
            "status_obra": "Concluído Campo" if status == "MEDIÇÃO CONCLUÍDA" else "Em Andamento",
            "prazo": prazo,
            "tempo_dias": tempo_dias,
            "atraso_dias": atraso_dias,
            "total_terceiros": total_terceiros,
            "previa_medicao": previa_medicao,
            "valor_medicao": valor_medicao,
            "valor_classe_l": val_l,
            "valor_classe_f": val_f,
            "itens_l_resumo": ", ".join(itens_l_resumo) if itens_l_resumo else "-",
            "itens_f_resumo": ", ".join(itens_f_resumo) if itens_f_resumo else "-",
            "lpu_itens": {
                "q_211": q_211, "q_212": q_212, "q_215": q_215, "q_113": q_113,
                "q_311": q_311, "q_317": q_317, "q_318": q_318, "q_315": q_315,
                "q_313": q_313, "q_314": q_314, "q_312": q_312
            }
        }
        records.append(record)

    print(f"Total de registros operacionais válidos extraídos: {len(records)}")

    # Extrair metadados e listas para filtros
    cidades = sorted(list(set(r["cidade"] for r in records if r["cidade"])))
    areas = sorted(list(set(r["area_tecnica"] for r in records if r["area_tecnica"])))
    status_list = sorted(list(set(r["status"] for r in records if r["status"])))
    prazos_list = ["NO PRAZO", "ATRASADO"]
    competencias = sorted(list(set(r["competencia"] for r in records if r["competencia"] and r["competencia"] != "NÃO INFORMADO")))
    anos = sorted(list(set(r["ano"] for r in records if r.get("ano") and r["ano"] != "NÃO INFORMADO")), reverse=True)
    
    # Competências e Anos de Entrega para o Fechamento de Terceiros
    competencias_entrega = sorted(list(set(r["competencia_entrega"] for r in records if r.get("competencia_entrega") and r["competencia_entrega"] != "NÃO INFORMADO")))
    anos_entrega = sorted(list(set(r["ano_entrega"] for r in records if r.get("ano_entrega") and r["ano_entrega"] != "NÃO INFORMADO")), reverse=True)

    # Competências e Anos de Medição
    comps_med_validas = [r["competencia_medicao"] for r in records if r.get("valor_medicao", 0) > 0 and r.get("competencia_medicao") and r["competencia_medicao"] not in ("SEM DATA (AJ)", "Sem Data")]
    
    def sort_comp_key(c):
        parts = c.split("/")
        if len(parts) == 2 and parts[1].isdigit():
            m_idx = MESES_PT.index(parts[0]) if parts[0] in MESES_PT else 0
            return (int(parts[1]), m_idx)
        return (0, 0)

    competencias_medicao = sorted(list(set(comps_med_validas)), key=sort_comp_key, reverse=True)
    if any(r.get("valor_medicao", 0) > 0 and r.get("competencia_medicao") in ("SEM DATA (AJ)", "Sem Data") for r in records):
        competencias_medicao.append("Sem Data")

    anos_medicao = sorted(list(set(r["ano_medicao"] for r in records if r.get("valor_medicao", 0) > 0 and r.get("ano_medicao") and r["ano_medicao"] != "SEM DATA")), reverse=True)

    terceiros_l = sorted(list(set(r["classe_l"] for r in records if r.get("classe_l"))))
    terceiros_f = sorted(list(set(r["classe_f"] for r in records if r.get("classe_f"))))
    todos_terceiros = sorted(list(set(terceiros_l + terceiros_f)))

    meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

    # Totais Globais
    tot_geral_terceiros = round(sum(r["total_terceiros"] for r in records), 2)
    tot_geral_previa = round(sum(r["previa_medicao"] for r in records), 2)
    tot_geral_l = round(sum(r["valor_classe_l"] for r in records), 2)
    tot_geral_f = round(sum(r["valor_classe_f"] for r in records), 2)

    # Totais Exclusivos de Medição (Coluna AK > 0 nos 3 status alvo)
    rec_med_alvo = [r for r in records if r.get("valor_medicao", 0) > 0 and r.get("status_medicao_grupo") in ("MEDIÇÃO ENVIADA", "FINALIZADO", "PEDIDO EMITIDO")]
    tot_med_geral = round(sum(r["valor_medicao"] for r in rec_med_alvo), 2)
    tot_med_enviada = round(sum(r["valor_medicao"] for r in rec_med_alvo if r["status_medicao_grupo"] == "MEDIÇÃO ENVIADA"), 2)
    tot_med_finalizado = round(sum(r["valor_medicao"] for r in rec_med_alvo if r["status_medicao_grupo"] == "FINALIZADO"), 2)
    tot_med_pedido = round(sum(r["valor_medicao"] for r in rec_med_alvo if r["status_medicao_grupo"] == "PEDIDO EMITIDO"), 2)

    qtd_med_geral = len(rec_med_alvo)
    qtd_med_enviada = len([r for r in rec_med_alvo if r["status_medicao_grupo"] == "MEDIÇÃO ENVIADA"])
    qtd_med_finalizado = len([r for r in rec_med_alvo if r["status_medicao_grupo"] == "FINALIZADO"])
    qtd_med_pedido = len([r for r in rec_med_alvo if r["status_medicao_grupo"] == "PEDIDO EMITIDO"])

    metadata = {
        "total_records": len(records),
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_file": source_desc,
        "cidades": cidades,
        "areas_tecnicas": areas,
        "status_list": status_list,
        "prazos": prazos_list,
        "competencias": competencias,
        "anos": anos,
        "competencias_entrega": competencias_entrega,
        "anos_entrega": anos_entrega,
        "competencias_medicao": competencias_medicao,
        "anos_medicao": anos_medicao,
        "terceiros_l": terceiros_l,
        "terceiros_f": terceiros_f,
        "todos_terceiros": todos_terceiros,
        "meses": meses,
        "financeiro": {
            "total_terceiros": tot_geral_terceiros,
            "total_previa_medicao": tot_geral_previa,
            "total_classe_l": tot_geral_l,
            "total_classe_f": tot_geral_f
        },
        "medicao": {
            "total_geral": tot_med_geral,
            "qtd_geral": qtd_med_geral,
            "total_medicao_enviada": tot_med_enviada,
            "qtd_medicao_enviada": qtd_med_enviada,
            "total_finalizado": tot_med_finalizado,
            "qtd_finalizado": qtd_med_finalizado,
            "total_pedido_emitido": tot_med_pedido,
            "qtd_pedido_emitido": qtd_med_pedido
        }
    }

    # Salvar sar_data.js
    output_js_path = os.path.join(base_dir, "sar_data.js")
    with open(output_js_path, "w", encoding="utf-8") as f:
        f.write(f"// Dados SAR JLE Telecom - Gerado em: {metadata['generated_at']}\n")
        f.write(f"// Fonte: {source_desc}\n")
        f.write("window.SAR_METADATA = " + json.dumps(metadata, ensure_ascii=False, indent=4) + ";\n\n")
        f.write("window.SAR_DATA = " + json.dumps(records, ensure_ascii=False, indent=2) + ";\n")

    print(f"Arquivo gerado com sucesso: {output_js_path}")

if __name__ == "__main__":
    main()
