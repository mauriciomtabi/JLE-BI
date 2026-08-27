#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_sar_operational_sheet.py
Gera a nova Planilha Operacional do SAR pré-formatada para o Google Sheets (Fran + Douglas),
migrando todos os 1.066 registros existentes e adicionando formatação corporativa e validações de dados.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    source_candidates = [
        r"\\10.121.21.252\matriz_rs\Claro\PROJETO F\Status Projeto F - Nodes, SAR - Nova Base.xlsx",
        os.path.join(base_dir, "sar_temp.xlsx"),
        os.path.join(base_dir, "sar_local.xlsx")
    ]
    
    source_file = None
    for c in source_candidates:
        if os.path.exists(c):
            source_file = c
            break
            
    if not source_file:
        print("ERRO: Nenhuma base de origem SAR encontrada!")
        return

    print(f"Lendo base original de: {source_file}")
    src_wb = openpyxl.load_workbook(source_file, data_only=True, read_only=True)
    src_ws = src_wb['Ext. MDU'] if 'Ext. MDU' in src_wb.sheetnames else src_wb.active

    # Criar novo workbook de destino
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "SAR Operacional"
    out_ws.views.sheetView[0].showGridLines = True

    # Paleta de Cores Corporativa
    FILL_HEADER_OP_SECTION = PatternFill(start_color="004F71", end_color="004F71", fill_type="solid")
    FILL_HEADER_OP_COL = PatternFill(start_color="006692", end_color="006692", fill_type="solid")
    
    FILL_HEADER_MED_SECTION = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")
    FILL_HEADER_MED_COL = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")

    FONT_SECTION_TITLE = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    FONT_COL_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    FONT_DATA = Font(name="Calibri", size=10, color="000000")
    
    BORDER_THIN = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    BORDER_HEADER = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='medium', color='002D42'),
        bottom=Side(style='medium', color='002D42')
    )

    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
    ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

    # Definição das Colunas
    # (Nome, Alinhamento, Largura Mínima, Bloco)
    columns_def = [
        # Bloco Operação (Fran) - Colunas 1 a 18 (A - R)
        ("CÓDIGO SAR", ALIGN_CENTER, 15, "OP"),
        ("CIDADE", ALIGN_LEFT, 18, "OP"),
        ("ÁREA TÉCNICA", ALIGN_CENTER, 14, "OP"),
        ("NODE", ALIGN_CENTER, 12, "OP"),
        ("SITE", ALIGN_CENTER, 12, "OP"),
        ("CONDOMÍNIO", ALIGN_LEFT, 24, "OP"),
        ("ENDEREÇO", ALIGN_LEFT, 32, "OP"),
        ("CAIXA MDU", ALIGN_CENTER, 16, "OP"),
        ("SERVIÇO / ESCOPO", ALIGN_LEFT, 35, "OP"),
        ("EXECUTOR LINHA (Classe L)", ALIGN_LEFT, 20, "OP"),
        ("EXECUTOR FUSÃO (Classe F)", ALIGN_LEFT, 20, "OP"),
        ("DATA DE ENTRADA", ALIGN_CENTER, 15, "OP"),
        ("INÍCIO EM", ALIGN_CENTER, 14, "OP"),
        ("PREVISÃO", ALIGN_CENTER, 14, "OP"),
        ("DATA DE ENTREGA", ALIGN_CENTER, 15, "OP"),
        ("RELATÓRIO PPT / FOTOS", ALIGN_CENTER, 20, "OP"),
        ("DATA ENVIO MEDIÇÃO", ALIGN_CENTER, 18, "OP"),
        ("STATUS OPERAÇÃO", ALIGN_CENTER, 18, "OP"),
        
        # Bloco Medição / Faturamento (Douglas) - Colunas 19 a 27 (S - AA)
        ("DATA MEDIÇÃO", ALIGN_CENTER, 15, "MED"),
        ("VALOR MEDIÇÃO (R$)", ALIGN_RIGHT, 18, "MED"),
        ("Nº WF", ALIGN_CENTER, 15, "MED"),
        ("STATUS WF", ALIGN_CENTER, 18, "MED"),
        ("Nº DO PEDIDO / CONTRATO", ALIGN_CENTER, 22, "MED"),
        ("STATUS GERAL SAR", ALIGN_CENTER, 20, "MED"),
        ("TEMPO (DIAS)", ALIGN_CENTER, 14, "MED"),
        ("PRAZO (SLA 3 DIAS)", ALIGN_CENTER, 18, "MED"),
        ("OBSERVAÇÕES", ALIGN_LEFT, 30, "MED")
    ]

    total_cols = len(columns_def)
    op_end_col = 18
    med_start_col = 19

    # 1. Escrever Linha 1: Super Headers (Seções)
    out_ws.row_dimensions[1].height = 28
    out_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=op_end_col)
    cell_op_title = out_ws.cell(row=1, column=1, value="ETAPA 1: OPERAÇÃO & EXECUÇÃO DE CAMPO (PREENCHIMENTO: FRAN)")
    cell_op_title.font = FONT_SECTION_TITLE
    cell_op_title.alignment = ALIGN_CENTER
    for col in range(1, op_end_col + 1):
        out_ws.cell(row=1, column=col).fill = FILL_HEADER_OP_SECTION
        out_ws.cell(row=1, column=col).border = BORDER_HEADER

    out_ws.merge_cells(start_row=1, start_column=med_start_col, end_row=1, end_column=total_cols)
    cell_med_title = out_ws.cell(row=1, column=med_start_col, value="ETAPA 2: MEDIÇÃO, WORKFLOW & FATURAMENTO (PREENCHIMENTO: DOUGLAS)")
    cell_med_title.font = FONT_SECTION_TITLE
    cell_med_title.alignment = ALIGN_CENTER
    for col in range(med_start_col, total_cols + 1):
        out_ws.cell(row=1, column=col).fill = FILL_HEADER_MED_SECTION
        out_ws.cell(row=1, column=col).border = BORDER_HEADER

    # 2. Escrever Linha 2: Cabeçalhos das Colunas
    out_ws.row_dimensions[2].height = 24
    for idx, (col_name, align, width, block) in enumerate(columns_def, start=1):
        c = out_ws.cell(row=2, column=idx, value=col_name)
        c.font = FONT_COL_HEADER
        c.alignment = ALIGN_CENTER
        c.fill = FILL_HEADER_OP_COL if block == "OP" else FILL_HEADER_MED_COL
        c.border = BORDER_HEADER
        col_letter = get_column_letter(idx)
        out_ws.column_dimensions[col_letter].width = width

    # Mapeamento dos índices da base original (linha 4)
    # Col B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, K=10, L=11, M=12, N=13, O=14, R=17, S=18, T=19, U=20, V=21, W=22, X=23, Y=24, Z=25, AA=26, AC=28, AD=29, AE=30, AF=31, AG=32, AH=33
    def clean_cell(val):
        if val is None: return ""
        if isinstance(val, (datetime.datetime, datetime.date)):
            return val
        s = str(val).strip()
        return "" if s.lower() in ("none", "null", "-", "n/a", "#n/a") else s

    records_count = 0
    current_out_row = 3

    for r_idx, row in enumerate(src_ws.iter_rows(min_row=5, values_only=True), start=5):
        if not row or len(row) < 2 or not row[1]:
            continue
            
        cod = clean_cell(row[1] if len(row) > 1 else "")
        if not cod:
            continue
            
        records_count += 1
        out_ws.row_dimensions[current_out_row].height = 20

        # Extrair campos da base original
        area_tecnica = clean_cell(row[2] if len(row) > 2 else "")
        node = clean_cell(row[3] if len(row) > 3 else "")
        site = clean_cell(row[4] if len(row) > 4 else "")
        cidade = clean_cell(row[5] if len(row) > 5 else "")
        condominio = clean_cell(row[6] if len(row) > 6 else "")
        endereco = clean_cell(row[7] if len(row) > 7 else "")
        caixa_mdu = clean_cell(row[8] if len(row) > 8 else "")
        status_obra = clean_cell(row[10] if len(row) > 10 else "")
        classe_l = clean_cell(row[11] if len(row) > 11 else "")
        classe_f = clean_cell(row[12] if len(row) > 12 else "")
        situacao = clean_cell(row[13] if len(row) > 13 else "")
        relat_foto = clean_cell(row[14] if len(row) > 14 else "")
        servico = clean_cell(row[17] if len(row) > 17 else "")
        envio_medicao = clean_cell(row[18] if len(row) > 18 else "")
        dt_entrada = clean_cell(row[19] if len(row) > 19 else "")
        dt_inicio = clean_cell(row[20] if len(row) > 20 else "")
        dt_previsao = clean_cell(row[21] if len(row) > 21 else "")
        dt_entrega = clean_cell(row[22] if len(row) > 22 else "")
        relat_ppt = clean_cell(row[23] if len(row) > 23 else "")
        
        # Colunas Medição / Douglas
        dt_medicao = clean_cell(row[24] if len(row) > 24 else "")
        num_wf = clean_cell(row[25] if len(row) > 25 else "")
        status_wf = clean_cell(row[26] if len(row) > 26 else "")
        status_geral = clean_cell(row[30] if len(row) > 30 else "") # Col AE
        if not status_geral:
            status_geral = clean_cell(row[28] if len(row) > 28 else "") # Col AC

        # Normalizar Status Geral
        sg_upper = str(status_geral).upper()
        if "MEDIC" in sg_upper and "CONCLU" in sg_upper:
            status_geral_norm = "MEDIÇÃO CONCLUÍDA"
        elif "WF APROV" in sg_upper or "FINALIZAD" in sg_upper or str(status_wf).upper() == "FINALIZADO":
            status_geral_norm = "MEDIÇÃO CONCLUÍDA"
        elif "SEM SINAL" in sg_upper:
            status_geral_norm = "SEM SINAL"
        elif "PARALISAD" in sg_upper:
            status_geral_norm = "PARALISADO"
        elif "CANCELAD" in sg_upper:
            status_geral_norm = "CANCELADO"
        elif "RELAT" in sg_upper or "AG. RELAT" in sg_upper or "AG RELAT" in sg_upper:
            status_geral_norm = "AG. RELATÓRIO"
        elif "AG. MEDI" in sg_upper or "AG MEDI" in sg_upper or "AG. APROV" in sg_upper:
            status_geral_norm = "AG. MEDIÇÃO"
        else:
            status_geral_norm = status_geral if status_geral else "AG. MEDIÇÃO"

        # Normalizar Status Operação
        so_upper = str(status_obra).upper()
        if "CONCLU" in so_upper:
            status_op_norm = "Concluído Campo"
        elif "CANCELAD" in so_upper:
            status_op_norm = "Cancelado"
        elif "SEM SINAL" in so_upper:
            status_op_norm = "Sem Sinal"
        elif "PARALISAD" in so_upper:
            status_op_norm = "Paralisado"
        else:
            status_op_norm = "Em Andamento"

        # Montar linha formatada de 27 colunas
        row_values = [
            cod,                    # A: CÓDIGO SAR
            cidade,                 # B: CIDADE
            area_tecnica,           # C: ÁREA TÉCNICA
            node,                   # D: NODE
            site,                   # E: SITE
            condominio,             # F: CONDOMÍNIO
            endereco,               # G: ENDEREÇO
            caixa_mdu,              # H: CAIXA MDU
            servico,                # I: SERVIÇO / ESCOPO
            classe_l,               # J: EXECUTOR LINHA
            classe_f,               # K: EXECUTOR FUSÃO
            dt_entrada,             # L: DATA DE ENTRADA
            dt_inicio,              # M: INÍCIO EM
            dt_previsao,            # N: PREVISÃO
            dt_entrega,             # O: DATA DE ENTREGA
            relat_ppt or relat_foto or "Não Precisa", # P: RELATÓRIO PPT / FOTOS
            envio_medicao,          # Q: DATA ENVIO MEDIÇÃO
            status_op_norm,         # R: STATUS OPERAÇÃO
            dt_medicao,             # S: DATA MEDIÇÃO
            "",                     # T: VALOR MEDIÇÃO (R$)
            num_wf,                 # U: Nº WF
            status_wf,              # V: STATUS WF
            "",                     # W: Nº DO PEDIDO / CONTRATO
            status_geral_norm,      # X: STATUS GERAL SAR
            f'=IF(AND(L{current_out_row}<>"", S{current_out_row}<>""), S{current_out_row}-L{current_out_row}, IF(L{current_out_row}<>"", TODAY()-L{current_out_row}, ""))', # Y: TEMPO (DIAS)
            f'=IF(Y{current_out_row}="", "", IF(Y{current_out_row}<=3, "NO PRAZO", "ATRASADO"))', # Z: PRAZO (SLA 3 DIAS)
            situacao                # AA: OBSERVAÇÕES
        ]

        for col_idx, val in enumerate(row_values, start=1):
            cell = out_ws.cell(row=current_out_row, column=col_idx, value=val)
            cell.font = FONT_DATA
            cell.border = BORDER_THIN
            align = columns_def[col_idx - 1][1]
            cell.alignment = align

            # Formatação de tipo
            if isinstance(val, (datetime.datetime, datetime.date)):
                cell.number_format = 'DD/MM/YYYY'
            elif col_idx == 20: # Valor Medição
                cell.number_format = 'R$ #,##0.00'
            elif col_idx == 25: # Tempo Dias
                cell.number_format = '#,##0'

        current_out_row += 1

    # 3. Adicionar Validações de Dados (Listas Suspensas)
    max_data_row = max(current_out_row + 1000, 3000)

    # DV: Relatório PPT / Fotos (Col P = 16)
    dv_ppt = DataValidation(type="list", formula1='"OK, Pendência, Não Precisa, -"', allow_blank=True)
    out_ws.add_data_validation(dv_ppt)
    dv_ppt.add(f"P3:P{max_data_row}")

    # DV: Status Operação (Col R = 18)
    dv_op = DataValidation(type="list", formula1='"Em Andamento, Concluído Campo, Sem Sinal, Paralisado, Cancelado"', allow_blank=True)
    out_ws.add_data_validation(dv_op)
    dv_op.add(f"R3:R{max_data_row}")

    # DV: Status WF (Col V = 22)
    dv_wf = DataValidation(type="list", formula1='"Aguardando WF, WF Aprovado, Pendência Claro, Cancelado, Finalizado"', allow_blank=True)
    out_ws.add_data_validation(dv_wf)
    dv_wf.add(f"V3:V{max_data_row}")

    # DV: Status Geral SAR (Col X = 24)
    dv_geral = DataValidation(type="list", formula1='"AG. MEDIÇÃO, MEDIÇÃO CONCLUÍDA, AG. RELATÓRIO, CANCELADO, PARALISADO, SEM SINAL"', allow_blank=True)
    out_ws.add_data_validation(dv_geral)
    dv_geral.add(f"X3:X{max_data_row}")

    # 4. Congelar Painéis (Linha 3 - Cabeçalhos fixos)
    out_ws.freeze_panes = "A3"
    
    # 5. Adicionar AutoFiltro
    out_ws.auto_filter.ref = f"A2:AA{current_out_row - 1}"

    # Salvar Arquivo
    out_filename = os.path.join(base_dir, "Planilha_Operacional_SAR_JLE.xlsx")
    out_wb.save(out_filename)
    print(f"\n[SUCESSO] Planilha Operacional do SAR gerada com {records_count} registros migrados!")
    print(f"Arquivo salvo em: {out_filename}")

if __name__ == '__main__':
    main()
