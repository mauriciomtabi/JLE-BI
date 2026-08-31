# -*- coding: utf-8 -*-
"""
Script ETL para extração de dados de Controle de Parcelamentos e Gestão Tributária
Lê a planilha da Controladoria e gera parcelamentos_data.js e parcelamentos_local.xlsx
Nota: A aba de impostos não parcelados é explicitamente ignorada conforme diretriz de negócio.
"""

import os
import sys
import json
import shutil
import datetime
import openpyxl

# Configuração de caminhos e diretórios
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOCAL_FALLBACK_FILE = os.path.join(SCRIPT_DIR, "parcelamentos_local.xlsx")
OUTPUT_JS_FILE = os.path.join(SCRIPT_DIR, "parcelamentos_data.js")

CANDIDATE_DIRS = [
    r"\\10.121.21.252\controladoria\Vitor\TRIBUTARIO",
    r"\\10.121.21.252\controladoria\Vitor",
    r"\\10.121.21.252\controladoria\TRIBUTARIO",
    r"\\10.121.21.252\controladoria"
]

def find_latest_network_file():
    """Localiza o arquivo mais recente de controle de parcelamentos na rede"""
    for c_dir in CANDIDATE_DIRS:
        if os.path.exists(c_dir):
            try:
                files = os.listdir(c_dir)
                matching = []
                for f in files:
                    if f.startswith("~$"):
                        continue
                    if f.lower().endswith((".xlsx", ".xls")) and ("parcelamento" in f.lower() or "tributario" in f.lower() or "parcelamentos" in f.lower()):
                        full_path = os.path.join(c_dir, f)
                        mtime = os.path.getmtime(full_path)
                        matching.append((mtime, full_path, f))
                
                if matching:
                    matching.sort(key=lambda x: x[0], reverse=True)
                    latest = matching[0][1]
                    print(f"[ETL] Arquivo de rede selecionado: {latest}")
                    return latest
            except Exception as e:
                print(f"[ETL] Aviso ao acessar {c_dir}: {e}")
    return None

def get_source_file():
    """Retorna o caminho do arquivo de origem (rede com fallback local)"""
    net_file = find_latest_network_file()
    if net_file and os.path.exists(net_file):
        try:
            shutil.copy2(net_file, LOCAL_FALLBACK_FILE)
            print(f"[ETL] Cache local atualizado com sucesso: {LOCAL_FALLBACK_FILE}")
            return net_file
        except Exception as e:
            print(f"[ETL] Erro ao copiar cache local (utilizando rede diretamente): {e}")
            return net_file
            
    if os.path.exists(LOCAL_FALLBACK_FILE):
        print(f"[ETL] Rede indisponível. Utilizando cache local de contingência: {LOCAL_FALLBACK_FILE}")
        return LOCAL_FALLBACK_FILE
        
    raise FileNotFoundError("[ETL] Nenhuma fonte de dados encontrada (rede e cache local indisponíveis).")

def parse_date(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()

def parse_float(v):
    if v is None or str(v).strip() in ("", "-", "N/D", "#REF!", "#N/A", "#VALOR!"):
        return 0.0
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).strip().replace("R$", "").replace(" ", "")
    # Formato pt-BR 1.234,56 -> 1234.56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return round(float(s), 2)
    except Exception:
        return 0.0

def process_workbook(file_path):
    print(f"[ETL] Carregando pasta de trabalho: {file_path}")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    data = {
        "metadata": {
            "source_file": os.path.basename(file_path),
            "generated_at": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            "reference_position": "Agosto/2026"
        },
        "overview": {},
        "acordos": [],
        "acompanhamento_mensal": [],
        "totais_gerais": {}
    }

    # 1. Identificar abas de parcelamento
    installment_sheet_names = [s for s in wb.sheetnames if "PARCELAMENTO" in s.upper() and s.strip().upper() != "PARCELAMENTOS"]
    
    total_divida_original = 0.0
    total_pago = 0.0
    total_saldo_devedor = 0.0
    total_juros_pagos = 0.0
    total_parcelas_compromisso_mensal = 0.0

    acordos_list = []

    for sname in installment_sheet_names:
        ws = wb[sname]
        clean_name = sname.strip()
        
        # Leitura da caixa de resumo (O4:R6)
        divida_orig = parse_float(ws.cell(4, 15).value)
        pago_acum = parse_float(ws.cell(4, 16).value)
        saldo_dev = parse_float(ws.cell(4, 17).value)
        juros_tot = parse_float(ws.cell(4, 18).value)

        # Ajuste de consistência se saldo devedor ou divida estiverem zerados na célula
        if divida_orig == 0 and ws.max_row > 4:
            divida_orig = sum(parse_float(ws.cell(r, 9).value) for r in range(4, ws.max_row + 1) if ws.cell(r, 9).value is not None)
        if saldo_dev == 0 and divida_orig > 0:
            saldo_dev = round(divida_orig - pago_acum, 2)

        # Determinar data inicial e identificador amigável
        label_acordo = clean_name.replace("PARCELAMENTO", "Parcelamento").strip()
        data_adesao = ""
        first_venc = ws.cell(4, 12).value
        if first_venc:
            data_adesao = parse_date(first_venc)
            
        # Extração dos Tributos de Origem (Colunas B a H)
        debitos_origem = []
        for r in range(4, ws.max_row + 1):
            cod = ws.cell(r, 2).value
            prin = ws.cell(r, 5).value
            saldo_col = ws.cell(r, 8).value
            if cod and prin is not None and str(cod).strip() != "":
                debitos_origem.append({
                    "id": f"{clean_name}_D{r}",
                    "codigo_receita": str(cod).strip(),
                    "apuracao": parse_date(ws.cell(r, 3).value),
                    "vencimento_original": parse_date(ws.cell(r, 4).value),
                    "valor_principal": parse_float(prin),
                    "valor_multa": parse_float(ws.cell(r, 6).value),
                    "valor_juros": parse_float(ws.cell(r, 7).value),
                    "saldo_consolidado": parse_float(saldo_col)
                })

        # Extração do Cronograma das Parcelas (Colunas I a M)
        parcelas_schedule = []
        parcelas_pagas_count = 0
        acum_pago_check = 0.0

        for r in range(4, ws.max_row + 1):
            v_parc = ws.cell(r, 9).value
            v_venc = ws.cell(r, 12).value
            
            if v_parc is not None and v_venc is not None:
                val_base = parse_float(v_parc)
                juros_selic = parse_float(ws.cell(r, 10).value)
                val_total = parse_float(ws.cell(r, 11).value)
                if val_total == 0:
                    val_total = round(val_base + juros_selic, 2)
                    
                venc_dt_str = parse_date(v_venc)
                saldo_remanescente = ws.cell(r, 13).value
                saldo_rem_float = parse_float(saldo_remanescente) if saldo_remanescente is not None else None

                p_num = len(parcelas_schedule) + 1

                is_paga = False
                if acum_pago_check + val_base <= pago_acum + 1.0:
                    is_paga = True
                    acum_pago_check += val_base
                    parcelas_pagas_count += 1
                elif juros_selic > 0:
                    is_paga = True
                    parcelas_pagas_count += 1

                parcelas_schedule.append({
                    "numero": p_num,
                    "vencimento": venc_dt_str,
                    "valor_base": val_base,
                    "juros_selic": juros_selic,
                    "valor_total": val_total,
                    "saldo_devedor_restante": saldo_rem_float,
                    "status": "Paga" if is_paga else "A Pagar"
                })

        total_parcelas_acordo = len(parcelas_schedule)
        parcelas_faltantes = max(0, total_parcelas_acordo - parcelas_pagas_count)
        
        # Parcela base mensal média
        parcela_mensal_base = parcelas_schedule[0]["valor_base"] if parcelas_schedule else 0.0
        if parcelas_faltantes > 0:
            total_parcelas_compromisso_mensal += parcela_mensal_base

        pct_amortizado = round((pago_acum / divida_orig * 100), 1) if divida_orig > 0 else 0.0

        acordo_obj = {
            "id": clean_name.lower().replace(" ", "_"),
            "nome": clean_name,
            "label": label_acordo,
            "data_adesao": data_adesao,
            "divida_original": divida_orig,
            "total_pago": pago_acum,
            "saldo_devedor": saldo_dev,
            "juros_totais_pagos": juros_tot,
            "pct_amortizado": pct_amortizado,
            "total_parcelas": total_parcelas_acordo,
            "parcelas_pagas": parcelas_pagas_count,
            "parcelas_faltantes": parcelas_faltantes,
            "valor_parcela_base": parcela_mensal_base,
            "debitos_origem": debitos_origem,
            "cronograma_parcelas": parcelas_schedule
        }

        acordos_list.append(acordo_obj)

        total_divida_original += divida_orig
        total_pago += pago_acum
        total_saldo_devedor += saldo_dev
        total_juros_pagos += juros_tot

    data["acordos"] = acordos_list

    # 2. Acompanhamento Mensal (Faturamento vs Tributos)
    sheet_acomp_list = [s for s in wb.sheetnames if "ACOMPANHAMENTO" in s.upper()]
    acompanhamento_mensal = []

    if sheet_acomp_list:
        ws_acomp = wb[sheet_acomp_list[0]]
        for c in range(3, ws_acomp.max_column + 1):
            dt_val = ws_acomp.cell(2, c).value
            if dt_val:
                mes_label = dt_val.strftime("%m/%Y") if isinstance(dt_val, (datetime.datetime, datetime.date)) else str(dt_val)
                fat = parse_float(ws_acomp.cell(3, c).value)
                iss = parse_float(ws_acomp.cell(4, c).value)
                pis = parse_float(ws_acomp.cell(5, c).value)
                cofins = parse_float(ws_acomp.cell(6, c).value)
                ir = parse_float(ws_acomp.cell(7, c).value)
                csll = parse_float(ws_acomp.cell(8, c).value)
                tot_impostos = parse_float(ws_acomp.cell(9, c).value)
                if tot_impostos == 0:
                    tot_impostos = round(iss + pis + cofins + ir + csll, 2)

                carga_pct = round((tot_impostos / fat * 100), 2) if fat > 0 else 0.0

                acompanhamento_mensal.append({
                    "mes": mes_label,
                    "faturamento": fat,
                    "iss": iss,
                    "pis": pis,
                    "cofins": cofins,
                    "ir": ir,
                    "csll": csll,
                    "total_impostos": tot_impostos,
                    "carga_tributaria_pct": carga_pct
                })

    data["acompanhamento_mensal"] = acompanhamento_mensal

    # 3. Projeção Mensal de Desembolso Futuro Consolidado (2026 até 2030)
    fluxo_futuro_map = {}
    for acordo in acordos_list:
        for p in acordo["cronograma_parcelas"]:
            if p["status"] == "A Pagar" and p["vencimento"]:
                try:
                    parts = p["vencimento"].split("/")
                    if len(parts) == 3:
                        mes_ano = f"{parts[1]}/{parts[2]}"
                        sort_key = f"{parts[2]}-{parts[1]}"
                    else:
                        mes_ano = p["vencimento"]
                        sort_key = p["vencimento"]
                except Exception:
                    mes_ano = p["vencimento"]
                    sort_key = p["vencimento"]

                if sort_key not in fluxo_futuro_map:
                    fluxo_futuro_map[sort_key] = {
                        "mes_ano": mes_ano,
                        "sort_key": sort_key,
                        "valor_total_projetado": 0.0,
                        "por_acordo": {}
                    }
                fluxo_futuro_map[sort_key]["valor_total_projetado"] += p["valor_base"]
                fluxo_futuro_map[sort_key]["por_acordo"][acordo["label"]] = round(
                    fluxo_futuro_map[sort_key]["por_acordo"].get(acordo["label"], 0.0) + p["valor_base"], 2
                )

    fluxo_futuro_sorted = sorted(fluxo_futuro_map.values(), key=lambda x: x["sort_key"])
    for item in fluxo_futuro_sorted:
        item["valor_total_projetado"] = round(item["valor_total_projetado"], 2)

    data["projecao_desembolso_futuro"] = fluxo_futuro_sorted

    # 4. Totais Gerais Consolidados (Apenas Parcelamentos Ativos)
    pct_quitado_total = round((total_pago / total_divida_original * 100), 1) if total_divida_original > 0 else 0.0

    data["totais_gerais"] = {
        "divida_original_parcelados": round(total_divida_original, 2),
        "total_pago_amortizado": round(total_pago, 2),
        "saldo_devedor_parcelado": round(total_saldo_devedor, 2),
        "juros_selic_pagos_parcelas": round(total_juros_pagos, 2),
        "compromisso_mensal_atual": round(total_parcelas_compromisso_mensal, 2),
        "pct_quitado_geral": pct_quitado_total,
        "total_acordos_ativos": len(acordos_list)
    }

    return data

def main():
    print("=" * 70)
    print("INICIANDO ETL: CONTROLE DE PARCELAMENTOS (ACORDOS VIGENTES)")
    print(f"Data/Hora: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 70)

    try:
        source_file = get_source_file()
        processed_data = process_workbook(source_file)

        # Gravação do arquivo JavaScript
        js_content = f"// Dados Consolidados de Parcelamentos Tributários JLE Telecom\n"
        js_content += f"// Gerado automaticamente pelo pipeline ETL em {processed_data['metadata']['generated_at']}\n\n"
        js_content += f"window.PARCELAMENTOS_DATA = {json.dumps(processed_data, indent=2, ensure_ascii=False)};\n"

        with open(OUTPUT_JS_FILE, "w", encoding="utf-8") as f:
            f.write(js_content)

        print(f"[ETL] Sucesso! Base compilada salva em: {OUTPUT_JS_FILE}")
        print("\n--- RESUMO EXECUTIVO DO PROCESSAMENTO ---")
        tot = processed_data["totais_gerais"]
        print(f"  • Dívida Consolidada Original: R$ {tot['divida_original_parcelados']:,.2f}")
        print(f"  • Total Já Amortizado:         R$ {tot['total_pago_amortizado']:,.2f} ({tot['pct_quitado_geral']:.1f}%)")
        print(f"  • Saldo Devedor Parcelado:     R$ {tot['saldo_devedor_parcelado']:,.2f} ({100-tot['pct_quitado_geral']:.1f}%)")
        print(f"  • Juros Selic Pagos:           R$ {tot['juros_selic_pagos_parcelas']:,.2f}")
        print(f"  • Compromisso Mensal Atual:    R$ {tot['compromisso_mensal_atual']:,.2f}/mês")
        print(f"  • Total de Acordos Ativos:     {tot['total_acordos_ativos']}")
        print("=" * 70)

    except Exception as e:
        print(f"[ETL] ERRO CRÍTICO no processamento: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
